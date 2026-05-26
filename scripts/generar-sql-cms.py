#!/usr/bin/env python3
"""Genera el SQL completo del CMS desde cms.xlsx: tablas + RLS + datos seed."""

import openpyxl
import os

wb = openpyxl.load_workbook("cms.xlsx", data_only=True)


def sql_esc(v):
    """Escapa string para SQL."""
    if v is None:
        return "NULL"
    s = str(v)
    return "'" + s.replace("'", "''") + "'"


def emit_table_and_policies(out, table_name, columns_sql):
    out.write(
        f"""
-- ============================================
-- TABLA {table_name}
-- ============================================
create table if not exists public.{table_name} (
{columns_sql}
);

alter table public.{table_name} enable row level security;

drop policy if exists "Anon leer {table_name}" on public.{table_name};
create policy "Anon leer {table_name}" on public.{table_name}
  for select to anon using (true);

drop policy if exists "Anon insertar {table_name}" on public.{table_name};
create policy "Anon insertar {table_name}" on public.{table_name}
  for insert to anon with check (true);

drop policy if exists "Anon actualizar {table_name}" on public.{table_name};
create policy "Anon actualizar {table_name}" on public.{table_name}
  for update to anon using (true) with check (true);

drop policy if exists "Anon eliminar {table_name}" on public.{table_name};
create policy "Anon eliminar {table_name}" on public.{table_name}
  for delete to anon using (true);

"""
    )


with open("scripts/cms_setup.sql", "w") as out:
    out.write(
        "-- =============================================================\n"
        "-- CMS COMPLETO — Valentina Centeno\n"
        "-- Pega TODO este bloque en Supabase → SQL Editor → New query → Run\n"
        "-- =============================================================\n"
    )

    # 1) cms_textos (key/value para ABOUT, FORMACION, INICIATIVAS)
    emit_table_and_policies(
        out,
        "cms_textos",
        """  id uuid primary key default gen_random_uuid(),
  seccion text not null,
  clave text not null,
  valor text,
  updated_at timestamptz not null default now(),
  unique (seccion, clave)""",
    )

    # Seed ABOUT
    ws = wb["ABOUT"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed ABOUT (solo si no existe ya en cms_textos)\n"
        "insert into public.cms_textos (seccion, clave, valor)\n"
        "select * from (values\n"
    )
    vals = []
    for r in rows[1:]:
        if r and r[0]:
            vals.append(f"  ('about', {sql_esc(r[0])}, {sql_esc(r[1])})")
    out.write(",\n".join(vals))
    out.write(
        "\n) as t(seccion, clave, valor)\n"
        "where not exists (select 1 from public.cms_textos where seccion='about');\n\n"
    )

    # Seed FORMACION
    ws = wb["FORMACION"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed FORMACION\n"
        "insert into public.cms_textos (seccion, clave, valor)\n"
        "select * from (values\n"
    )
    vals = []
    for r in rows[1:]:
        if r and r[0]:
            vals.append(f"  ('formacion', {sql_esc(r[0])}, {sql_esc(r[1])})")
    out.write(",\n".join(vals))
    out.write(
        "\n) as t(seccion, clave, valor)\n"
        "where not exists (select 1 from public.cms_textos where seccion='formacion');\n\n"
    )

    # Seed INICIATIVAS (1 fila: sectionTitle, description)
    ws = wb["INICIATIVAS"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed INICIATIVAS\n"
        "insert into public.cms_textos (seccion, clave, valor)\n"
        "select * from (values\n"
    )
    headers = rows[0]
    vals = []
    if len(rows) > 1:
        r = rows[1]
        for i, h in enumerate(headers):
            if h and i < len(r) and r[i] is not None:
                vals.append(f"  ('iniciativas', {sql_esc(str(h))}, {sql_esc(r[i])})")
    out.write(",\n".join(vals) if vals else "  ('iniciativas', 'sectionTitle', '')")
    out.write(
        "\n) as t(seccion, clave, valor)\n"
        "where not exists (select 1 from public.cms_textos where seccion='iniciativas');\n"
    )

    # 2) cms_leyes
    emit_table_and_policies(
        out,
        "cms_leyes",
        """  id uuid primary key default gen_random_uuid(),
  orden int not null default 0,
  title_top text,
  title text,
  img text,
  descripcion text,
  full_text text,
  section_title text,
  section_subtitle text,
  activo boolean not null default true,
  updated_at timestamptz not null default now()""",
    )

    ws = wb["LEYES"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed LEYES\n"
        "insert into public.cms_leyes (orden, title_top, title, img, descripcion, full_text, section_title, section_subtitle, activo)\n"
        "select * from (values\n"
    )
    vals = []
    for r in rows[1:]:
        if r and r[0]:
            orden = int(r[0]) if r[0] else 0
            vals.append(
                f"  ({orden}, {sql_esc(r[1])}, {sql_esc(r[2])}, {sql_esc(r[3])}, "
                f"{sql_esc(r[4])}, {sql_esc(r[5])}, {sql_esc(r[6])}, {sql_esc(r[7])}, true)"
            )
    out.write(",\n".join(vals))
    out.write(
        "\n) as t(orden, title_top, title, img, descripcion, full_text, section_title, section_subtitle, activo)\n"
        "where not exists (select 1 from public.cms_leyes);\n"
    )

    # 3) cms_logros
    emit_table_and_policies(
        out,
        "cms_logros",
        """  id uuid primary key default gen_random_uuid(),
  orden int not null default 0,
  icon_key text,
  title text,
  body text,
  image text,
  section_title text,
  activo boolean not null default true,
  updated_at timestamptz not null default now()""",
    )

    ws = wb["BLOGROS"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed LOGROS\n"
        "insert into public.cms_logros (orden, icon_key, title, body, image, section_title, activo)\n"
        "select * from (values\n"
    )
    vals = []
    seq = 0
    for r in rows[1:]:
        if r and r[2]:  # require title
            seq += 1
            vals.append(
                f"  ({seq}, {sql_esc(r[1])}, {sql_esc(r[2])}, {sql_esc(r[3])}, "
                f"{sql_esc(r[4])}, {sql_esc(r[5])}, true)"
            )
    out.write(",\n".join(vals))
    out.write(
        "\n) as t(orden, icon_key, title, body, image, section_title, activo)\n"
        "where not exists (select 1 from public.cms_logros);\n"
    )

    # 4) cms_agenda_internacional
    emit_table_and_policies(
        out,
        "cms_agenda_internacional",
        """  id uuid primary key default gen_random_uuid(),
  title text,
  tag text,
  subtitle text,
  description text,
  bullets text,
  quote text,
  image text,
  updated_at timestamptz not null default now()""",
    )

    ws = wb["AGENDAINTER"]
    rows = list(ws.iter_rows(values_only=True))
    out.write("-- Seed AGENDAINTER\n")
    if len(rows) > 1:
        r = rows[1]
        out.write(
            "insert into public.cms_agenda_internacional (title, tag, subtitle, description, bullets, quote, image)\n"
            f"select {sql_esc(r[1])}, {sql_esc(r[2])}, {sql_esc(r[3])}, {sql_esc(r[4])}, {sql_esc(r[5])}, {sql_esc(r[6])}, {sql_esc(r[7])}\n"
            "where not exists (select 1 from public.cms_agenda_internacional);\n"
        )

    # 5) cms_biografia
    emit_table_and_policies(
        out,
        "cms_biografia",
        """  id uuid primary key default gen_random_uuid(),
  orden int not null default 0,
  seccion text not null,
  identificador text,
  titulo text,
  descripcion text,
  imagen text,
  extra text,
  updated_at timestamptz not null default now()""",
    )

    ws = wb["BIOGRAFIA"]
    rows = list(ws.iter_rows(values_only=True))
    out.write(
        "-- Seed BIOGRAFIA\n"
        "insert into public.cms_biografia (orden, seccion, identificador, titulo, descripcion, imagen, extra)\n"
        "select * from (values\n"
    )
    vals = []
    seq = 0
    for r in rows[1:]:
        if r and r[0]:
            seq += 1
            vals.append(
                f"  ({seq}, {sql_esc(r[0])}, {sql_esc(r[1])}, {sql_esc(r[2])}, "
                f"{sql_esc(r[3])}, {sql_esc(r[4])}, {sql_esc(r[5])})"
            )
    out.write(",\n".join(vals))
    out.write(
        "\n) as t(orden, seccion, identificador, titulo, descripcion, imagen, extra)\n"
        "where not exists (select 1 from public.cms_biografia);\n"
    )

print("Listo: scripts/cms_setup.sql")
print(f"Tamaño: {os.path.getsize('scripts/cms_setup.sql')} bytes")
